# -- coding: utf-8 --'
from  openpyxl import *
import sqlite3
from unidecode import unidecode


pasta='d:\\rif\\'

def abrirtabela( xcursor, nometabela, campos=[]):
    try:
        sql = 'DROP TABLE IF EXISTS ' + nometabela
        xcursor.execute(sql)
        xcursor.connection.commit()
    except:
        print('Tabela ', nometabela, ' não pode ser excluída: ', sql)

    try:

        sql = 'CREATE TABLE IF NOT EXISTS ' + nometabela + '('
        sql = sql + ' text, '.join(campos)
        sql = sql + ' text)'
        xcursor.execute(sql)
        xcursor.connection.commit()
    except:
        print('Tabela ', nometabela, ' não pode ser criada: ', sql)
    try:
        sql = 'DELETE FROM ' + nometabela
        xcursor.execute(sql)
        xcursor.connection.commit()
    except:
        print('Tabela ', nometabela, 'não pode ser apagada: ', sql)
    return sql

def buscacampos(filename = pasta+'Comunicacoes.xlsx'):
    wb = load_workbook(filename)
    ws = wb.get_active_sheet()

    c=[]
    row = ws[1]
    for cell in row:
        c.append(cell.value)
    return c
#for col in ws.columns:
#    for cell in col:
#        print(cell.value)
def executaCursor(cursor, sql):
    try:
        r = cursor.execute(sql)
    except sqlite3.DataError as e:
        print("DataError")
        print(e)

    except sqlite3.InternalError as e:
        print("InternalError")
        print(e)

    except sqlite3.IntegrityError as e:
        print("IntegrityError")
        print(e)

    except sqlite3.OperationalError as e:
        print("OperationalError")
        print(e)

    except sqlite3.NotSupportedError as e:
        print("NotSupportedError")
        print(e)

    except sqlite3.ProgrammingError as e:
        print("ProgrammingError")
        print(e)

    except:
        print('Não foi possível conectar o BD SQLite')
def celulasvalores(lv,linha):
    for cell in linha:

        try:
            v = unidecode(str(cell.value))
        except:
            v='#'
        #v.replace('"', "").replace('&', 'e').replace('–', '-').replace(u'\u201c', '').replace(u'\u201d', '')
        lv.append(v)
    return(lv)

def buscavalores(cursor,filename,nometabela):
    wb = load_workbook(filename)
    ws = wb.get_active_sheet()
    campos=buscacampos(filename)
    abrirtabela(cursor,nometabela,campos)
    lv=[]
    i=0
    print(len(campos))
    for row in ws.rows:
        i+=1
        if i ==1:

            continue
        else:
            lv=celulasvalores(lv,row)
            sql_insere_inicio = u"insert into " + nometabela + " ("
            sql_insere_fim = u" values (\""
            sql_insere_inicio += ', '.join(campos) + ')'
            sql_insere_fim += '", "'.join(lv) + '")'
            print(sql_insere_inicio + sql_insere_fim)
            executaCursor(cursor,sql_insere_inicio + sql_insere_fim)
        lv.clear()

    return lv

try:
    # conn1 = sqlite3.connect("rif_"+codinome+".db")  # ou use :memory: para botÃ¡-lo na memÃ³ria RAM
    conn1 = sqlite3.connect(":memory:")  # ou use :memory: para botÃ¡-lo na memÃ³ria RAM
    cursor1 = conn1.cursor()
except sqlite3.DataError as e:
    print("DataError")
    print(e)

except sqlite3.InternalError as e:
    print("InternalError")
    print(e)

except sqlite3.IntegrityError as e:
    print("IntegrityError")
    print(e)

except sqlite3.OperationalError as e:
    print("OperationalError")
    print(e)

except sqlite3.NotSupportedError as e:
    print("NotSupportedError")
    print(e)

except sqlite3.ProgrammingError as e:
    print("ProgrammingError")
    print(e)

except:
    print('Não foi possível conectar o BD SQLite')


buscavalores(cursor1,pasta+'Comunicacoes.xlsx','COMUNICACOES')
buscavalores(cursor1,pasta+'Envolvidos.xlsx','ENVOLVIDOS')
buscavalores(cursor1,pasta+'Ocorrencias.xlsx','OCORRENCIAS')
buscavalores(cursor1,pasta+'Pedido.xlsx','PEDIDO')

#sql=u'''select * from COMUNICACOES'''
sql=u'''SELECT  ENVOLVIDOS.Indexador as INDEXADOR, ENVOLVIDOS.tipoEnvolvido AS tipoEnvolvido, ENVOLVIDOS.cpfCnpjEnvolvido AS CNPJCPF,'''
sql +=''' ENVOLVIDOS.nomeEnvolvido AS nomeEnvolvido, COMUNICACOES.Data_do_Recebimento AS Data_do_Recebimento, COMUNICACOES.'''
sql +=''' Data_da_operacao AS Data_da_operacao, COMUNICACOES.nomeComunicante AS nomeComunicante,'''
sql +=''' COMUNICACOES.informacoesAdicionais AS informacoesAdicionais, COMUNICACOES.CampoA AS CampoA, COMUNICACOES.CampoB AS CampoB,'''
sql +=''' COMUNICACOES.CampoC AS CampoC, COMUNICACOES.CampoD AS CampoD, COMUNICACOES.CampoE AS CampoE,'''
sql +=''' PEDIDO.Empresa_Nome AS NOME_PEDIDO, PEDIDO.JUSTIFICATIVA AS JUSTIFICATIVA_PEDIDO'''
sql +=''' FROM ENVOLVIDOS LEFT OUTER JOIN COMUNICACOES ON ENVOLVIDOS.Indexador = COMUNICACOES.Indexador '''
sql +=''' LEFT OUTER  JOIN PEDIDO ON REPLACE(REPLACE(REPLACE(ENVOLVIDOS.cpfCnpjEnvolvido,'.',''),'-',''),'/','') = REPLACE(REPLACE(REPLACE(PEDIDO.cpfCnpjEnvolvido,'.',''),'-',''),'/','') '''
saida = Workbook(write_only=True)
ws = saida.create_sheet()
try:


    r= cursor1.execute(sql)
    r= cursor1.fetchall()

    print([i[0] for i in cursor1.description])
    ws.append([i[0] for i in cursor1.description])
    for l in r:
        ws.append(l)
        #print(l)

except sqlite3.DataError as e:
    print("DataError")
    print(e)

except sqlite3.InternalError as e:
    print("InternalError")
    print(e)

except sqlite3.IntegrityError as e:
    print("IntegrityError")
    print(e)

except sqlite3.OperationalError as e:
    print("OperationalError")
    print(e)

except sqlite3.NotSupportedError as e:
    print("NotSupportedError")
    print(e)

except sqlite3.ProgrammingError as e:
    print("ProgrammingError")
    print(e)
#except:
#    print('Não foi possível executar a consulta ao BD SQLite')

saida.save(pasta+'consolidado.xlsx')
